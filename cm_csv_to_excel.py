"""Convert CSV files from output directory to Excel workbook with Power Query connections."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font
except ImportError as e:
    print(f"Error: Required library not installed: {e}")
    print("Install with: pip install pandas openpyxl")
    exit(1)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


def get_csv_files(output_dir: Path) -> List[Path]:
    """Get all CSV files from the output directory.
    
    Args:
        output_dir: Directory containing CSV files
        
    Returns:
        List of CSV file paths
    """
    csv_files = sorted(output_dir.glob("*.csv"))
    logger.info(f"Found {len(csv_files)} CSV file(s) in {output_dir}")
    return csv_files


def sanitize_sheet_name(name: str) -> str:
    """Sanitize filename to valid Excel sheet name.
    
    Excel sheet names:
    - Max 31 characters
    - Cannot contain: [ ] : * ? / \
    - Cannot be empty
    
    Args:
        name: Original filename/name
        
    Returns:
        Sanitized name valid for Excel
    """
    # Remove extension if present
    name = Path(name).stem
    
    # Replace invalid characters
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    for char in invalid_chars:
        name = name.replace(char, '_')
    
    # Truncate to 31 characters
    if len(name) > 31:
        name = name[:31]
    
    # Ensure not empty
    if not name:
        name = "Sheet1"
    
    return name


def format_excel_sheet(workbook, sheet_name: str) -> None:
    """Apply formatting to an Excel sheet.
    
    Args:
        workbook: OpenPyXL workbook object
        sheet_name: Name of the sheet to format
    """
    if sheet_name not in workbook.sheetnames:
        return
    
    sheet = workbook[sheet_name]
    
    # 1. Autofit column widths (approximate)
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        
        # Set width (add some padding, max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Freeze top row
    sheet.freeze_panes = 'A2'
    
    # 3. Enable filter on header row
    if sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    
    # 4. Style specific header columns (if they exist)
    target_headers = ['logged_domain_id', 'logged_domain_name', 'logged_as_user']
    
    # Define styles
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    dark_blue_font = Font(color='003366', bold=True)
    
    # Get header row (first row)
    header_row = sheet[1]
    
    # Find and format target headers (case-insensitive search)
    for header_cell in header_row:
        if header_cell.value is None:
            continue
            
        header_value = str(header_cell.value).strip()
        
        # Check if this column matches one of our target headers (case-insensitive)
        if header_value.lower() in [h.lower() for h in target_headers]:
            # Apply formatting to header cell
            header_cell.fill = yellow_fill
            header_cell.font = dark_blue_font
            
            column_letter = get_column_letter(header_cell.column)
            logger.debug(f"  Formatted header '{header_value}' ({column_letter}1) - yellow background, dark blue text")


def csv_to_excel_sheets(csv_files: List[Path], excel_file: Path) -> None:
    """Convert CSV files to Excel sheets with formatting.
    
    Args:
        csv_files: List of CSV file paths
        excel_file: Path to Excel output file
    """
    if not csv_files:
        logger.warning("No CSV files found to process")
        return
    
    # Create Excel writer
    logger.info(f"Creating Excel file: {excel_file}")
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for csv_file in csv_files:
            try:
                logger.info(f"Processing: {csv_file.name}")
                
                # Read CSV file
                df = pd.read_csv(csv_file, encoding='utf-8')
                
                # Get sanitized sheet name
                sheet_name = sanitize_sheet_name(csv_file.name)
                
                # Write to Excel sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                logger.info(f"  Added sheet '{sheet_name}' with {len(df)} rows, {len(df.columns)} columns")
                
            except Exception as e:
                logger.error(f"Failed to process {csv_file.name}: {e}")
                continue
    
    # Apply formatting to all sheets
    logger.info("Applying formatting to sheets...")
    workbook = load_workbook(excel_file)
    
    for sheet_name in workbook.sheetnames:
        logger.info(f"  Formatting sheet: {sheet_name}")
        format_excel_sheet(workbook, sheet_name)
    
    workbook.save(excel_file)
    workbook.close()
    
    logger.info(f"Successfully created Excel file: {excel_file}")


def add_power_query_connections(csv_files: List[Path], excel_file: Path) -> None:
    """Add Power Query connections to Excel file (Windows only, requires Excel COM).
    
    This function creates refreshable connections to CSV files that can be used
    to create Power Query-enabled sheets in Excel.
    
    Args:
        csv_files: List of CSV file paths
        excel_file: Path to Excel file
    """
    try:
        import win32com.client
        
        logger.info("Attempting to add Power Query connections using COM automation...")
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            workbook = excel.Workbooks.Open(str(excel_file.absolute()))
            
            for csv_file in csv_files:
                try:
                    sheet_name = sanitize_sheet_name(csv_file.name)
                    csv_path = str(csv_file.absolute())
                    
                    logger.info(f"  Creating Power Query query for: {csv_file.name}")
                    
                    # Build M formula for Power Query
                    try:
                        preview_df = pd.read_csv(csv_file, nrows=0)
                        num_columns = len(preview_df.columns)
                    except Exception:
                        num_columns = 0
                    
                    m_formula = (
                        "let\n"
                        f"    Source = Csv.Document(File.Contents(\"{csv_path}\"),[Delimiter=\",\", Columns={num_columns}, Encoding=65001, QuoteStyle=QuoteStyle.Csv]),\n"
                        "    #\"Promoted Headers\" = Table.PromoteHeaders(Source, [PromoteAllScalars=true])\n"
                        "in\n"
                        "    #\"Promoted Headers\""
                    )
                    
                    # Ensure Queries collection exists (Excel 2016+)
                    created_query = None
                    try:
                        queries = workbook.Queries
                        query_name = f"PQ_{sheet_name}"
                        
                        # Delete existing query with same name
                        try:
                            for q in list(queries):
                                if getattr(q, "Name", None) == query_name:
                                    q.Delete()
                                    break
                        except Exception:
                            pass
                        
                        created_query = queries.Add(query_name, m_formula)
                        logger.info(f"    Created Power Query: {query_name}")
                    except Exception as qe:
                        logger.warning(f"    Could not create Power Query for {csv_file.name}: {qe}")
                    
                    # Do not create legacy Workbook Connections; only create PQ Queries
                    # Users can load each query to a table or connection from Excel UI as needed
                    
                except Exception as e:
                    logger.warning(f"Could not add Power Query for {csv_file.name}: {e}")
                    continue
            
            workbook.Save()
            workbook.Close()
            logger.info("Power Query connections processing completed")
            
        finally:
            excel.Quit()
            
    except ImportError:
        logger.info("win32com not available - skipping Power Query connections")
        logger.info("Data has been loaded as Excel sheets. Power Query connections can be added manually in Excel.")
    except Exception as e:
        logger.warning(f"Could not add Power Query connections: {e}")
        logger.info("Data has been loaded as Excel sheets. Power Query connections can be added manually in Excel.")


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Convert CSV files from output directory to Excel workbook"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output"),
        help="Directory containing CSV files (default: output)"
    )
    parser.add_argument(
        "--excel-file",
        type=Path,
        default=Path("cm_reports.xlsx"),
        help="Output Excel file path (default: cm_reports.xlsx)"
    )
    parser.add_argument(
        "--add-power-query",
        action="store_true",
        help="Attempt to add Power Query connections (Windows only, requires Excel)"
    )
    
    args = parser.parse_args()
    
    # Check output directory exists
    if not args.output_dir.exists():
        logger.error(f"Output directory does not exist: {args.output_dir}")
        return 1
    
    # Get CSV files
    csv_files = get_csv_files(args.output_dir)
    
    if not csv_files:
        logger.warning(f"No CSV files found in {args.output_dir}")
        return 1
    
    try:
        # Convert CSV files to Excel sheets
        csv_to_excel_sheets(csv_files, args.excel_file)
        
        # Always add Power Query connections after creating sheets
        add_power_query_connections(csv_files, args.excel_file)
        
        logger.info(f"Successfully created: {args.excel_file}")
        return 0
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    exit(main())
